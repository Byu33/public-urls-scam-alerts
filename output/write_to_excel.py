from __future__ import annotations

import csv
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import PatternFill


REPO_ROOT = Path(__file__).resolve().parent.parent
AGENTS_DIR = REPO_ROOT / "agents"
OUTPUT_DIR = REPO_ROOT / "sharepoint_output"
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from local.scam_filters import parse_date  # noqa: E402
from local.supabase_client import fetch_all, get_supabase_client  # noqa: E402


def _read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not fieldnames:
        keys: list[str] = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        fieldnames = keys or ["message"]

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        if rows:
            writer.writerows(rows)


def _tier_rows(prefix: str, counts: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {"source": prefix, "alert_tier": tier, "count": int(counts.get(tier, 0) or 0)}
        for tier in ("CRITICAL", "ALERT", "WATCH")
    ]


def _recent_rows(rows: list[dict[str, Any]], field: str, days: int) -> list[dict[str, Any]]:
    cutoff = datetime.now(timezone.utc).timestamp() - days * 24 * 60 * 60
    kept = []
    for row in rows:
        parsed = parse_date(row.get(field))
        if parsed and parsed.timestamp() >= cutoff:
            kept.append(row)
    return kept


def _top_category(categories: dict[str, Any]) -> str | None:
    if not categories:
        return None
    return max(categories.items(), key=lambda item: int(item[1] or 0))[0]


def _apply_row_colors(workbook_path: Path) -> None:
    from openpyxl import load_workbook

    fills = {
        "Government Impersonation": PatternFill("solid", fgColor="F4CCCC"),
        "Identity Theft": PatternFill("solid", fgColor="FCE5CD"),
        "Payment Fraud": PatternFill("solid", fgColor="FFF2CC"),
        "Elder Fraud": PatternFill("solid", fgColor="D9D2E9"),
        "arrest": PatternFill("solid", fgColor="D9EAD3"),
        "warning": PatternFill("solid", fgColor="F4CCCC"),
        "advisory": PatternFill("solid", fgColor="FFF2CC"),
        "report": PatternFill("solid", fgColor="CFE2F3"),
        "HIGH": PatternFill("solid", fgColor="EA9999"),
        "MEDIUM": PatternFill("solid", fgColor="F9CB9C"),
        "LOW": PatternFill("solid", fgColor="B6D7A8"),
    }

    wb = load_workbook(workbook_path)
    for sheet_name, column_name in (
        ("Chicago Crime", "scam_category"),
        ("NYC Crime", "scam_category"),
        ("LA Crime", "scam_category"),
        ("News and Press Releases", "sentiment"),
        ("Intelligence Summary", "combined_risk"),
    ):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        if column_name not in headers:
            continue
        idx = headers.index(column_name) + 1
        for row in ws.iter_rows(min_row=2):
            fill = fills.get(row[idx - 1].value)
            if fill:
                for cell in row:
                    cell.fill = fill
    wb.save(workbook_path)


def _local_outputs() -> dict[str, Any]:
    summary = _read_json(AGENTS_DIR / "local_intelligence_summary.json")
    try:
        client = get_supabase_client()
        crime_rows = fetch_all(client, "local_crime_reports")
        news_rows = fetch_all(client, "local_news_mentions")
    except Exception as exc:
        print(f"WARNING: local Supabase export skipped: {exc}")
        crime_rows = []
        news_rows = []

    crime_7d = _recent_rows(crime_rows, "report_date", 7)
    news_14d = [
        row for row in _recent_rows(news_rows, "published_at", 14)
        if int(row.get("keyword_match_count") or 0) >= 2
    ]

    def crime_for(city: str) -> list[dict[str, Any]]:
        return sorted(
            [row for row in crime_7d if row.get("city") == city],
            key=lambda row: (str(row.get("scam_category") or ""), str(row.get("report_date") or "")),
            reverse=True,
        )

    chicago = crime_for("Chicago")
    nyc = crime_for("New York")
    la = crime_for("Los Angeles")
    news_14d = sorted(news_14d, key=lambda row: str(row.get("published_at") or ""), reverse=True)

    city_summary_rows = []
    cross_source_rows = []
    for city, data in (summary.get("cities") or {}).items():
        matches = data.get("cross_source_matches", []) or []
        city_summary_rows.append(
            {
                "city": city,
                "combined_risk": data.get("combined_risk"),
                "crime_records_7days": data.get("crime_records_7days"),
                "crime_deviation": data.get("crime_deviation"),
                "crime_spike": data.get("crime_spike"),
                "top_crime_scam_category": _top_category(data.get("crime_by_scam_category") or {}),
                "news_warnings": data.get("news_warnings"),
                "news_arrests": data.get("news_arrests"),
                "keyword_overlap_count": len(data.get("keyword_overlap_crime_news") or []),
                "bbb_alerts_count": len(data.get("bbb_alerts") or []),
                "cfpb_alerts_count": len(data.get("cfpb_alerts") or []),
                "cross_source_matches_count": len(matches),
                "risk_reason": data.get("risk_reason"),
            }
        )
        for match in matches:
            cross_source_rows.append(
                {
                    "city": city,
                    "scam_type": match.get("scam_type"),
                    "bbb_tier": match.get("bbb_tier"),
                    "cfpb_tier": match.get("cfpb_tier"),
                    "local_crime_count": match.get("local_crime_count"),
                    "local_news_count": match.get("local_news_count"),
                    "confidence_level": match.get("confidence_level"),
                    "description": match.get("description"),
                }
            )

    crime_chicago_cols = ["report_date", "offense_type", "scam_category", "description", "location_description", "community_area"]
    crime_nyc_cols = ["report_date", "offense_type", "scam_category", "description", "borough"]
    crime_la_cols = ["report_date", "offense_type", "scam_category", "description", "division", "location_description"]
    news_cols = [
        "published_at",
        "city",
        "source",
        "sentiment",
        "scam_category",
        "keyword_match_count",
        "scam_keywords_found",
        "headline",
        "summary",
    ]
    summary_cols = [
        "city",
        "combined_risk",
        "crime_records_7days",
        "crime_deviation",
        "crime_spike",
        "top_crime_scam_category",
        "news_warnings",
        "news_arrests",
        "keyword_overlap_count",
        "bbb_alerts_count",
        "cfpb_alerts_count",
        "cross_source_matches_count",
        "risk_reason",
    ]
    cross_cols = [
        "city",
        "scam_type",
        "bbb_tier",
        "cfpb_tier",
        "local_crime_count",
        "local_news_count",
        "confidence_level",
        "description",
    ]

    _write_csv(OUTPUT_DIR / "Local_Crime_Chicago.csv", chicago, crime_chicago_cols)
    _write_csv(OUTPUT_DIR / "Local_Crime_NYC.csv", nyc, crime_nyc_cols)
    _write_csv(OUTPUT_DIR / "Local_Crime_LA.csv", la, crime_la_cols)
    _write_csv(OUTPUT_DIR / "Local_News_Scam_Mentions.csv", news_14d, news_cols)
    _write_csv(OUTPUT_DIR / "Local_Intelligence_Summary.csv", city_summary_rows, summary_cols)
    _write_csv(OUTPUT_DIR / "Local_Cross_Source_Signals.csv", cross_source_rows, cross_cols)

    workbook_path = OUTPUT_DIR / "Local_Law_Enforcement.xlsx"
    sheets = {
        "Chicago Crime": (chicago, crime_chicago_cols),
        "NYC Crime": (nyc, crime_nyc_cols),
        "LA Crime": (la, crime_la_cols),
        "News and Press Releases": (news_14d, news_cols),
        "Intelligence Summary": (city_summary_rows, summary_cols),
        "Cross Source Signals": (cross_source_rows, cross_cols),
    }
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for sheet_name, (rows, columns) in sheets.items():
            pd.DataFrame(rows, columns=columns).to_excel(writer, sheet_name=sheet_name, index=False)

    _apply_row_colors(workbook_path)
    row_counts = {sheet_name: len(rows) for sheet_name, (rows, _) in sheets.items()}
    return {
        "excel_file": str(workbook_path.relative_to(REPO_ROOT)),
        "sheet_row_counts": row_counts,
        "csv_files": [
            "sharepoint_output/Local_Crime_Chicago.csv",
            "sharepoint_output/Local_Crime_NYC.csv",
            "sharepoint_output/Local_Crime_LA.csv",
            "sharepoint_output/Local_News_Scam_Mentions.csv",
            "sharepoint_output/Local_Intelligence_Summary.csv",
            "sharepoint_output/Local_Cross_Source_Signals.csv",
        ],
    }


def _briefing_markdown(
    builder: dict[str, Any],
    verifier: dict[str, Any],
    quality: dict[str, Any],
    output_status: str,
) -> str:
    data_summary = {
        "BBB records ingested": builder.get("data_summary", {}).get("bbb_records_ingested", 0),
        "CFPB records ingested": builder.get("data_summary", {}).get("cfpb_records_ingested", 0),
        "Local crime records": builder.get("data_summary", {}).get("local_crime_records", 0),
        "BBB anomalies": verifier.get("anomaly_summary", {}).get("bbb_anomalies_by_tier", {}),
        "CFPB anomalies": verifier.get("anomaly_summary", {}).get("cfpb_anomalies_by_tier", {}),
        "Cross source signals": quality.get("cross_source_signals", {}).get("count", 0),
    }
    lines = [
        "# Weekly Scam Intelligence Briefing",
        "",
        f"Generated: {datetime.now(timezone.utc).isoformat()}",
        "",
        "## Phase Status",
        f"- Builder: {builder.get('overall_status', 'UNKNOWN')}",
        f"- Verifier: {verifier.get('overall_status', 'UNKNOWN')}",
        f"- Data Quality: {quality.get('overall_status', 'UNKNOWN')}",
        f"- Output: {output_status}",
        "",
        "## Data Summary",
    ]
    for key, value in data_summary.items():
        lines.append(f"- {key}: {value}")

    warnings = []
    for report in (builder, verifier, quality):
        warnings.extend(report.get("warnings", []) or [])
        warnings.extend(report.get("issues", []) or [])
    if warnings:
        lines.extend(["", "## Warnings and Follow-up"])
        lines.extend(f"- {warning}" for warning in warnings)

    return "\n".join(lines) + "\n"


def write_outputs() -> dict[str, Any]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    builder = _read_json(AGENTS_DIR / "builder_report.json")
    verifier = _read_json(AGENTS_DIR / "verifier_report.json")
    quality = _read_json(AGENTS_DIR / "quality_report.json")

    table_counts = [
        {"table": table, "row_count": count}
        for table, count in sorted((verifier.get("table_counts") or {}).items())
    ]
    _write_csv(OUTPUT_DIR / "Supabase_Table_Counts.csv", table_counts, ["table", "row_count"])

    _write_csv(
        OUTPUT_DIR / "BBB_Anomalies.csv",
        _tier_rows("BBB", verifier.get("anomaly_summary", {}).get("bbb_anomalies_by_tier", {})),
        ["source", "alert_tier", "count"],
    )
    _write_csv(
        OUTPUT_DIR / "CFPB_Anomalies.csv",
        _tier_rows("CFPB", verifier.get("anomaly_summary", {}).get("cfpb_anomalies_by_tier", {})),
        ["source", "alert_tier", "count"],
    )
    _write_csv(
        OUTPUT_DIR / "Cross_Source_Signals.csv",
        quality.get("cross_source_signals", {}).get("signals", []),
        ["signal_type", "description", "count"],
    )
    _write_csv(
        OUTPUT_DIR / "Pipeline_Phase_Status.csv",
        [
            {"phase": "Builder", "status": builder.get("overall_status", "UNKNOWN")},
            {"phase": "Verifier", "status": verifier.get("overall_status", "UNKNOWN")},
            {"phase": "Data Quality", "status": quality.get("overall_status", "UNKNOWN")},
            {"phase": "Output", "status": "PASS"},
        ],
        ["phase", "status"],
    )

    briefing_path = OUTPUT_DIR / "Weekly_Briefing.md"
    briefing_path.write_text(_briefing_markdown(builder, verifier, quality, "PASS"), encoding="utf-8")

    workbook_path = OUTPUT_DIR / "Scam_Intelligence_Briefing.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(table_counts).to_excel(writer, sheet_name="Table Counts", index=False)
        pd.DataFrame(_tier_rows("BBB", verifier.get("anomaly_summary", {}).get("bbb_anomalies_by_tier", {}))).to_excel(
            writer, sheet_name="BBB Anomalies", index=False
        )
        pd.DataFrame(_tier_rows("CFPB", verifier.get("anomaly_summary", {}).get("cfpb_anomalies_by_tier", {}))).to_excel(
            writer, sheet_name="CFPB Anomalies", index=False
        )
        pd.DataFrame(quality.get("narrative_samples", [])).to_excel(writer, sheet_name="Narrative Samples", index=False)
        pd.DataFrame(quality.get("cross_source_signals", {}).get("signals", [])).to_excel(
            writer, sheet_name="Cross Signals", index=False
        )

    local_output = _local_outputs()
    files = sorted(str(path.relative_to(REPO_ROOT)) for path in OUTPUT_DIR.iterdir() if path.is_file())
    result = {
        "overall_status": "PASS",
        "files": files,
        "excel_file": str(workbook_path.relative_to(REPO_ROOT)),
        "local_excel_file": local_output.get("excel_file"),
        "local_sheet_row_counts": local_output.get("sheet_row_counts", {}),
        "local_csv_files": local_output.get("csv_files", []),
        "briefing_file": str(briefing_path.relative_to(REPO_ROOT)),
    }
    print(json.dumps(result, indent=2))
    return result


def main() -> dict[str, Any]:
    return write_outputs()


if __name__ == "__main__":
    main()
